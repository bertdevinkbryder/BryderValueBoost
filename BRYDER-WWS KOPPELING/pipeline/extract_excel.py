"""Extraheert eenheden-informatie uit het Excel-bestand naar CSV."""

import csv
import openpyxl
from pathlib import Path


def extract_eenheden(excel_path: Path, output_path: Path):
    """
    Leest het Excel-bestand en schrijft eenheden.csv.
    Filtert bedrijfsruimten (bouwnummer=None) eruit.
    """
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Lees headers
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Zoek kolom-indices
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[h.strip()] = i

    fieldnames = [
        "eenheid_id",
        "bouwnummer",
        "straat",
        "huisnummer",
        "huisletter",
        "huisnummertoevoeging",
        "postcode",
        "plaats",
        "type_overzicht",
        "cluster_naam",
        "pand_id",
        "bag_verblijfsobject",
        "bag_nummeraanduiding",
        "woningwaarderingstelsel",
        "pandsoort",
        "bouwjaar",
        "klimaatbeheersing",
        "woz_waarde",
        "woz_peildatum",
        "energielabel",
        "energieprestatie_soort",
        "energieprestatie_waarde",
        "energieprestatie_begindatum",
        "energieprestatie_einddatum",
        "monument",
        "gebruiksoppervlakte",
    ]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    count = 0

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()

        for row in ws.iter_rows(min_row=2, values_only=True):
            bouwnummer = row[col_map.get("bouwnummer", -1)]
            if bouwnummer is None:
                continue  # Skip bedrijfsruimten

            oge = row[col_map.get("OGEnummer", 0)]
            type_ovz = row[col_map.get("type overzicht", -1)] or ""

            # Bepaal pandsoort op basis van type
            if type_ovz.startswith("B1"):
                pandsoort = "EGW"  # Eengezinswoning (laagbouw)
            else:
                pandsoort = "MGW"  # Meergezinswoning (B2/B3 appartementen)

            writer.writerow({
                "eenheid_id": oge,
                "bouwnummer": int(bouwnummer),
                "straat": row[col_map.get("straat", -1)] or "",
                "huisnummer": row[col_map.get("Huisnummer", -1)] or "",
                "huisletter": row[col_map.get("Huis Letter", -1)] or "",
                "huisnummertoevoeging": row[col_map.get("Huisnummertoevoeging", -1)] or "",
                "postcode": row[col_map.get("Postcode", -1)] or "",
                "plaats": row[col_map.get("Plaats", -1)] or "",
                "type_overzicht": type_ovz,
                "cluster_naam": row[col_map.get("Cluster Naam Algemeen", -1)] or "",
                "pand_id": row[col_map.get("Pand_id", -1)] or "",
                "bag_verblijfsobject": row[col_map.get("BAG", -1)] or "",
                "bag_nummeraanduiding": row[col_map.get("BAG Nummeraanduiding", -1)] or "",
                "woningwaarderingstelsel": "ZEL",
                "pandsoort": pandsoort,
                "bouwjaar": "",  # Handmatig invullen
                "klimaatbeheersing": "",  # Handmatig invullen
                "woz_waarde": "",
                "woz_peildatum": "",
                "energielabel": "",
                "energieprestatie_soort": "",
                "energieprestatie_waarde": "",
                "energieprestatie_begindatum": "",
                "energieprestatie_einddatum": "",
                "monument": "",
                "gebruiksoppervlakte": "",
            })
            count += 1

    print(f"Eenheden CSV geschreven: {output_path} ({count} eenheden)")
    return count
